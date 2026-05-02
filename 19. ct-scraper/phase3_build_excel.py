# phase3_build_excel.py
# Export database → Excel 9 tab, linked by Docket Number
#
# Jalankan: python phase3_build_excel.py

import sqlite3, os, logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH    = "scraper.db"
OUTPUT_DIR = "data/final"
LOG_FILE   = "logs/phase3.log"

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ])
log = logging.getLogger(__name__)

# ── STYLES ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_FONT  = Font(size=9)
ROW_ALIGN = Alignment(vertical="top", wrap_text=False)
ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")
WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
LNK_FONT  = Font(size=9, color="0563C1", underline="single")
THIN      = Side(style="thin", color="D0D0D0")
BORDER    = Border(left=THIN, right=THIN, bottom=THIN)

def style_ws(ws, nrows, ncols, link_col_indices=None):
    link_col_indices = link_col_indices or []
    ws.row_dimensions[1].height = 28
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN

    for r in range(2, nrows + 2):
        fill = ALT_FILL if r % 2 == 0 else WHT_FILL
        for c in range(1, ncols + 1):
            cell        = ws.cell(row=r, column=c)
            cell.fill   = fill
            cell.border = BORDER
            if c in link_col_indices and cell.value and str(cell.value).startswith("http"):
                cell.font      = LNK_FONT
                cell.hyperlink = str(cell.value)
            else:
                cell.font      = ROW_FONT
                cell.alignment = ROW_ALIGN

def auto_col_width(ws, max_w=60):
    for col in ws.columns:
        letter  = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[letter].width = min(max_len + 2, max_w)

def write_tab(wb, name, df, link_cols=None):
    log.info(f"  Building tab: {name} ({len(df):,} rows)...")
    ws = wb.create_sheet(title=name)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{min(len(df)+1, 1000000)}"

    # Header
    for i, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=i, value=col)

    # Data
    link_indices = []
    if link_cols:
        link_indices = [list(df.columns).index(c) + 1 for c in link_cols if c in df.columns]

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
        if r_idx % 20000 == 0:
            log.info(f"    {name}: {r_idx:,} rows written...")

    style_ws(ws, len(df), len(df.columns), link_indices)
    auto_col_width(ws)
    log.info(f"    {name}: done ✅")

# ── DATA QUERIES ──────────────────────────────────────────────────────────────
def q1_case_info(conn):
    return pd.read_sql("""
        SELECT
            ac_number                                        AS "Docket Number",
            TRIM(REPLACE(REPLACE(title,  char(160),' '),'&nbsp;',' '))  AS "Case Title",
            TRIM(REPLACE(REPLACE(status, char(160),' '),'&nbsp;',' '))  AS "Status",
            crn                                              AS "CRN"
        FROM case_information
        ORDER BY ac_number
    """, conn)

def q2_appeal_case_info(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number        AS "Docket Number",
            a.date_filed        AS "Date Filed",
            a.response_due_date AS "Response to Docket Due Date",
            a.appeal_by         AS "Appeal By",
            a.disposition_method AS "Disposition Method",
            a.argued_date       AS "Argued Date",
            a.disposition_date  AS "Disposition Date",
            a.submitted_briefs_date AS "Submitted on Briefs Date",
            a.cite              AS "Cite",
            a.panel             AS "Panel",
            a.petitions_certification AS "Petition(s) For Certification"
        FROM appeal_case_info a
        JOIN case_information ci ON a.crn = ci.crn
        ORDER BY ci.ac_number
    """, conn)

def q3_cross_appeal(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number AS "Docket Number",
            c.raw_text   AS "Cross Appeal / Amended Appeal Info"
        FROM cross_appeal c
        JOIN case_information ci ON c.crn = ci.crn
        WHERE c.raw_text != ''
        ORDER BY ci.ac_number
    """, conn)

def q4_trial_court(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number        AS "Docket Number",
            t.tc_docket_number  AS "Trial Court Docket Number",
            t.judgment_for      AS "Judgment For",
            t.court             AS "Court",
            t.trial_judge       AS "Trial Judge(s)",
            t.judgment_date     AS "Judgment Date",
            t.f1                AS "Additional Field 1",
            t.f2                AS "Additional Field 2",
            t.f3                AS "Additional Field 3"
        FROM trial_court_info t
        JOIN case_information ci ON t.crn = ci.crn
        ORDER BY ci.ac_number
    """, conn)

def q5_party_attorney(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number     AS "Docket Number",
            p.party_name     AS "Party Name",
            p.party_class    AS "Party Class (Appellant/Appellee/etc)",
            p.juris_number   AS "Juris Number",
            p.juris_name     AS "Attorney Name",
            p.attorney_info  AS "Attorney / Firm Info"
        FROM party_attorney p
        JOIN case_information ci ON p.crn = ci.crn
        ORDER BY ci.ac_number, p.party_class, p.juris_number
    """, conn)

def q6_transcripts(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number  AS "Docket Number",
            t.col1        AS "Party Name",
            t.col2        AS "Order Date",
            t.col3        AS "Due Date",
            t.col4        AS "Filed Date",
            t.col5        AS "Pages",
            t.link_url    AS "Document Link"
        FROM transcripts_exhibits t
        JOIN case_information ci ON t.crn = ci.crn
        ORDER BY ci.ac_number
    """, conn)

def q7_prelim_papers(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number  AS "Docket Number",
            p.col1        AS "Party Name",
            p.col2        AS "Due Date",
            p.col3        AS "Filed Date",
            p.col4        AS "Received Date",
            p.col5        AS "Sent Date",
            p.link_url    AS "Document Link"
        FROM preliminary_papers p
        JOIN case_information ci ON p.crn = ci.crn
        ORDER BY ci.ac_number
    """, conn)

def q8_briefs(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number  AS "Docket Number",
            b.col1        AS "Party Name",
            b.col2        AS "Brief Type",
            b.col3        AS "Due Date",
            b.col4        AS "Filed Date",
            b.col5        AS "Received Date",
            b.link_url    AS "Document Link"
        FROM briefs_record b
        JOIN case_information ci ON b.crn = ci.crn
        ORDER BY ci.ac_number
    """, conn)

def q9_activities(conn):
    return pd.read_sql("""
        SELECT
            ci.ac_number    AS "Docket Number",
            a.activity_date AS "Date Filed",
            -- Activity field sekarang berisi: ACTIVITY | Description | By: Initiated
            CASE
                WHEN INSTR(a.activity, ' | ') > 0
                THEN SUBSTR(a.activity, 1, INSTR(a.activity, ' | ')-1)
                ELSE a.activity
            END             AS "Activity",
            CASE
                WHEN INSTR(a.activity, ' | ') > 0
                THEN TRIM(SUBSTR(a.activity, INSTR(a.activity, ' | ')+3))
                ELSE ''
            END             AS "Description / Details",
            a.link_url      AS "Document Link"
        FROM case_activity a
        JOIN case_information ci ON a.crn = ci.crn
        ORDER BY ci.ac_number, a.activity_date
    """, conn)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)

    total = pd.read_sql("SELECT COUNT(*) AS n FROM case_information", conn).iloc[0]["n"]
    log.info("=" * 60)
    log.info(f"PHASE 3 — Building Excel. Cases in DB: {total:,}")
    log.info("=" * 60)

    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(OUTPUT_DIR, f"CT_Appellate_Cases_{ts}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Hapus sheet default

    log.info("Loading and writing tabs...")
    write_tab(wb, "1_Case Information",       q1_case_info(conn))
    write_tab(wb, "2_Appeal Case Info",        q2_appeal_case_info(conn))
    write_tab(wb, "3_Cross Appeal",            q3_cross_appeal(conn))
    write_tab(wb, "4_Trial Court Info",        q4_trial_court(conn))
    write_tab(wb, "5_Party Attorney",          q5_party_attorney(conn))
    write_tab(wb, "6_Transcripts Exhibits",    q6_transcripts(conn),
              link_cols=["Document Link"])
    write_tab(wb, "7_Preliminary Papers",      q7_prelim_papers(conn),
              link_cols=["Document Link"])
    write_tab(wb, "8_Briefs Prepared Record",  q8_briefs(conn),
              link_cols=["Document Link"])
    write_tab(wb, "9_Case Activity",           q9_activities(conn),
              link_cols=["Document Link"])

    log.info(f"Saving to {out_path}...")
    wb.save(out_path)
    conn.close()

    size_mb = os.path.getsize(out_path) / (1024 * 1024)
    log.info(f"✅ File saved: {out_path}")
    log.info(f"   Size: {size_mb:.1f} MB")
    if size_mb > 250:
        log.warning("File > 250MB — beri tahu klien untuk buka di mesin RAM 8GB+")
    log.info("PHASE 3 DONE. Next: python phase4_validate.py")

if __name__ == "__main__":
    main()
