# phase4_validate.py — Pre-delivery QC check
# Jalankan: python phase4_validate.py

import sqlite3, os, glob
import pandas as pd
import openpyxl

DB_PATH    = "scraper.db"
OUTPUT_DIR = "data/final"

def main():
    conn = sqlite3.connect(DB_PATH)
    print("\n" + "═" * 60)
    print("  VALIDATION REPORT")
    print("═" * 60)

    # ── 1. Database status ──────────────────────────────────────────
    total   = conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
    done    = conn.execute("SELECT COUNT(*) FROM cases WHERE scrape_status='done'").fetchone()[0]
    failed  = conn.execute("SELECT COUNT(*) FROM cases WHERE scrape_status='failed'").fetchone()[0]
    sealed  = conn.execute("SELECT COUNT(*) FROM cases WHERE scrape_status='sealed'").fetchone()[0]
    pending = conn.execute("SELECT COUNT(*) FROM cases WHERE scrape_status='pending'").fetchone()[0]
    pct     = done / total * 100 if total else 0

    print(f"\n📊 DATABASE STATUS:")
    print(f"  Total cases    : {total:,}")
    print(f"  ✅ Done        : {done:,}  ({pct:.1f}%)")
    print(f"  ❌ Failed      : {failed:,}")
    print(f"  🔒 Sealed      : {sealed:,}")
    print(f"  ⏳ Pending     : {pending:,}")

    if pending > 0:
        print(f"\n  ⚠️  MASIH ADA {pending:,} PENDING — run phase2 lagi!")

    # ── 2. Row counts per table ─────────────────────────────────────
    tables = [
        ("case_information",    "Tab 1: Case Information"),
        ("appeal_case_info",    "Tab 2: Appeal Case Info"),
        ("cross_appeal",        "Tab 3: Cross Appeal"),
        ("trial_court_info",    "Tab 4: Trial Court Info"),
        ("party_attorney",      "Tab 5: Party/Attorney"),
        ("transcripts_exhibits","Tab 6: Transcripts"),
        ("preliminary_papers",  "Tab 7: Prelim Papers"),
        ("briefs_record",       "Tab 8: Briefs"),
        ("case_activity",       "Tab 9: Activities"),
    ]
    print(f"\n📋 ROW COUNTS PER TABLE:")
    for tbl, label in tables:
        cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        status = "✅" if cnt > 0 else "⚠️ "
        print(f"  {status} {label}: {cnt:,}")

    # ── 3. Null check di case_information ──────────────────────────
    sample = pd.read_sql("SELECT * FROM case_information LIMIT 5000", conn)
    print(f"\n🔍 NULL CHECK (case_information, sample 5K):")
    for col in ["ac_number", "title", "status"]:
        null_pct = sample[col].isna().sum() / len(sample) * 100
        flag = "⚠️ " if null_pct > 50 else "✅"
        print(f"  {flag} {col}: {null_pct:.0f}% null")

    # ── 4. Party tab sanity check ───────────────────────────────────
    party_sample = pd.read_sql("""
        SELECT juris_number, COUNT(*) AS cnt FROM party_attorney
        WHERE juris_number != '' AND juris_number != 'Self-Represented'
        GROUP BY juris_number HAVING cnt > 1 LIMIT 3
    """, conn)
    print(f"\n👥 PARTY TAB — Juris duplicates check (should be from different cases):")
    print(party_sample.to_string(index=False) if not party_sample.empty else "  No duplicates ✅")

    # ── 5. Sample data display ──────────────────────────────────────
    print(f"\n🧪 SAMPLE CASES (first 3):")
    s = pd.read_sql("SELECT ac_number, title, status FROM case_information LIMIT 3", conn)
    print(s.to_string(index=False))

    # ── 6. Excel file check ─────────────────────────────────────────
    excel_files = glob.glob(os.path.join(OUTPUT_DIR, "*.xlsx"))
    if excel_files:
        print(f"\n📁 EXCEL FILES:")
        for f in sorted(excel_files):
            size_mb = os.path.getsize(f) / (1024 * 1024)
            print(f"  {os.path.basename(f)}: {size_mb:.1f} MB")
            try:
                wb = openpyxl.load_workbook(f, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = ws.max_row
                    cols = ws.max_column
                    flag = "✅" if rows > 1 else "⚠️ "
                    print(f"    {flag} {sheet}: {rows:,} rows × {cols} cols")
                wb.close()
            except Exception as e:
                print(f"  ❌ Cannot open: {e}")
    else:
        print(f"\n⚠️  No Excel files found in {OUTPUT_DIR}")
        print("    Run phase3_build_excel.py first")

    conn.close()
    print("\n" + "═" * 60)
    print("  Validation complete.")
    if pending == 0 and done > 0:
        print("  ✅ READY TO DELIVER")
    else:
        print(f"  ⚠️  {pending:,} cases still pending")
    print("═" * 60 + "\n")

if __name__ == "__main__":
    main()
